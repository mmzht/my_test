import os

import openpyxl  # 只能读写xlsx文件
import pandas as pd
import xlrd  # xlrd不支持最新的xlsx文件，只支持xls文件
import xlwt  # xlrd不支持最新的xlsx文件，只支持xls

'''
filename = r'D:\python\data\excelfile.xlsx'  # 带路径的文件名
# filename = r'D:\python\data\excelfile.xls'  # 带路径的文件名
savefile = r'D:\python\data\savefile.xlsx'  # 结果保存文件

# 只读取xlsx文件#
if os.path.splitext(filename)[1].lower() == '.xsx':  # 兼容大小写，根据后缀判断是否xlsx文件
    try:
        openpyxl.load_workbook(filename)  # 打开excel文件
    except:
        print('excel文件打开失败')
# 只能写入xlsx文件
workbook = openpyxl.Workbook()
sheet = workbook.active  # 获取活动的sheet
sheet['A1'] = '对A1单元格写的数据'
workbook.save('new.xlsx')  # 保存文件，只能保存为xlsx格式

# 只读取xls文件#
if os.path.splitext(filename)[1].lower() == '.xls':  # 根据后缀判断是否lsx文件
    book = xlrd.open_workbook(filename)  # xlrd不支持最新的xlsx文件，只支持xls
    sheet1 = book.sheets()[0]  # 表格的第一个sheet
else:
    print('仅支持xls格式的文件')
print('表格有数据的总行数', sheet1.nrows)
print('表格有数据的总列数', sheet1.ncols)
print('第3行值', sheet1.row_values(2))
print('第3列值', sheet1.col_values(2))
cell_3_3 = sheet1.cell(2, 2).value
print('第3行第3列的单元格的值：', cell_3_3)

# 只能写入xls文件
workbook = xlwt.Workbook()  # 打开excel程序
worksheet = workbook.add_sheet('new-sheet')  # 增加一个名称为new-sheet的sheet
worksheet.write(0, 0, '对A1（0，0）单元格写的数据')
worksheet.add_sheet('sheet_name')  # 增加sheet
workbook.save('excelfile.xls')  # 保存文件，只能保存为xls格式
'''

'''
read_excel(读excel函数的参数解释：
io, 文件名，带路径
sheet_name = 0,  是一个表格的字典,默认sheetname为0，返回多表使用sheetname=[0,1]，若sheetname=None是返回全表
header = 0, 指定作为列名的行，默认0，即取第一行，数据为列名行以下的数据；若数据不含列名，则设定header = None
names = None, 指定一到多列的名字，传入一个list数据
index_col = None, 指定某列为索引列
usecols = None, 读取指定的列，usecols = range(1, 3) 使用 [1, 3) 行，不包括第 3 行； usecols = [4, 7]使用 4和7行
skiprows = None, 省略指定行数的数据
skipfooter = 0, 从尾部开始略去行的数据

to_excel（写excel参数说明
excel_writer：文件路径，不存在的会自动生成
sheet_name =“Sheet1”：指定写的表
columns = None：指定输出某些列， columns = [“Name”, “Number”]
header = True：是否保存头行列名
index = True：是否保存索引列
startcol = 0：起始行
merge_cells = True：是否合并单元格
encoding = None：指定编码，常用utf - 8
float_format = None：浮点数保存的格式，默认保存为字符串；float_format =’ % .2f’  # 保存为浮点数，保留2位小数
engine = None：保存格式，指定io.excel.xlsx.writer、 io.excel.xls.writer、io.excel.xlsm.writer.
'''

import datetime

import pandas as pd
from pandas.tseries import offsets

filename = r'D:\python\data\excelfile.xlsx'  # 带路径的文件名
save_file = r'D:\python\data\savefile.xlsx'  # 结果保存文件


# use_cols_list = [1, 3, 5, 6, 8]  #usecols=use_cols_list, header=None, index_col=None
df = pd.read_excel(filename, sheet_name=0, usecols=range(7), nrows=8, index_col=None)
print(df.head())
# print(df.index.values, df.columns.values)  # 列序号/表头
# print(df.columns, df.index, df.shape)  # 获取行列序号，及行数和列数
# print(df.iat[2, 4])  # 访问某一个单元格，或df.at[2, '第3列']需要索引名称
# print(df.iloc[:2, :4])  # 访问一块数据df.loc[:2, '第3列']需要索引名称

# temp = df['第4列'].replace([44, 74, 94], 440)  # 替换某几个内容，原表没变，需要赋值回去才会变
# temp = df['第4列'].replace({44: 440, 74: 740, 94: 940})  # 按字典一一替换，原表没变
# temp = df.sort_values(by=['第4列', '第5列'], ascending={False, True})  # 排序，原表没变
# temp = df.drop('第2列', axis=1)  # 删除列，原表没变？
# temp = df.drop(index=[4, 5])  # 删除行，原表没变？
# temp = df.drop(columns=['第2列', '第5列'])  # 删除列，原表没变？
# temp = df.drop([1, 3], axis=0)  # 删除行
# temp = df[df['第4列'] > 100]  # 选择大于100的值
# temp = df['第4列'].value_counts()  # 统计每个数值出现次数
# temp = df['第4列'].value_counts(normalize=True)  # 统计每个数值出现比例
# temp = df['第4列'].unique()  # 获取唯一数值（不重复值列表）
# temp = df['第4列'].isin([34, 74])  # 数值查找
# df.insert(column="插入新列", value=[1, 100, 6, 4, 8, 7, 8, 9, ], loc=5)  # 插入列：column列名,value数据，插入位置loc
# temp = df.T  # 转置，原表不变
# temp = df.stack()  # 转换成单列，原表不变
# temp = df.melt()  # 转换成单列，原表不变
# temp = df['第4列'].apply(lambda x: x - 100)  # 对列进行元素进行函数运算，原表值不变
# temp = df.applymap(lambda x: x - 50)   # 对所有元素进行函数运算，原表值不变
# temp = df * 2
# temp = df['第4列'] > df['第7列'] #每个元素一一比较
# temp = df.count(axis=1)  #计算非空元素个数
# temp = df['第4列'].count()
# temp = df.sum(axis=1)
# temp = df.quantile(0.25)  # 四分之一分位的大小
# temp = df.corr()  # 相关性
# temp = df['日期2']-df['日期1'] #必须是完整的日期加时间才能相减
# df['时间1'] = df['时间1'].astype(str)  # pandas时间不能直接计算需要转换为字符串，再to_datetime
# df['时间1'] = pd.to_datetime(df['时间1'])
# # df['时间1'] = df['时间1'] - datetime.timedelta(0, 15600)  # timedelta，第一个是天，第二个是秒
# df['时间1'] = df['时间1'] - offsets.Day(0)-offsets.Hour(4)-offsets.Minute(20)  # offsets时间偏移
print(df.head())

# df.to_excel(save_file, sheet_name='第1张表', index=False, header=True)  # 覆盖保存，不方便保存多张表

writer = pd.ExcelWriter(save_file)  # 覆盖保存，方便保存多张表
df.to_excel(writer, sheet_name='表格1')
df.to_excel(writer, sheet_name='表格2')
writer.close()
